from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any


EMAIL_RE = re.compile(r"[\w.+-]+@[\w.-]+\.[A-Za-z]{2,}")


@dataclass
class Professor:
    name: str
    email: str = ""
    university: str = ""
    department: str = ""
    website: str = ""
    lab: str = ""
    research_focus: str = ""
    past_work: str = ""
    recent_papers: str = ""
    alignment: str = ""
    additional_details: str = ""
    openings_page: str = ""
    priority: str = ""
    status: str = ""
    raw: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        return {
            "name": self.name,
            "email": self.email,
            "university": self.university,
            "department": self.department,
            "website": self.website,
            "lab": self.lab,
            "research_focus": self.research_focus,
            "past_work": self.past_work,
            "recent_papers": self.recent_papers,
            "alignment": self.alignment,
            "additional_details": self.additional_details,
            "openings_page": self.openings_page,
            "priority": self.priority,
            "status": self.status,
            "raw": self.raw,
        }


def load_professors(path: Path) -> list[Professor]:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return _load_json(path)
    if suffix == ".csv":
        return _load_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx(path)
    raise ValueError(f"Unsupported professor file type: {path.suffix}")


def _load_json(path: Path) -> list[Professor]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        rows = payload.get("professors") or payload.get("data") or payload.get("rows")
        if rows is None:
            rows = [payload]
    elif isinstance(payload, list):
        rows = payload
    else:
        raise ValueError("Professor JSON must be a list or an object containing professors/data/rows.")
    return [_row_to_professor(row) for row in rows if isinstance(row, dict)]


def _load_csv(path: Path) -> list[Professor]:
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return [_row_to_professor(row) for row in csv.DictReader(handle)]


def _load_xlsx(path: Path) -> list[Professor]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is missing. Run: pip install -r requirements.txt") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    professors: list[Professor] = []
    for row_values in rows[1:]:
        row = {
            headers[index]: _cell_to_text(value)
            for index, value in enumerate(row_values)
            if index < len(headers) and headers[index]
        }
        if any(row.values()):
            professors.append(_row_to_professor(row))
    return professors


def _row_to_professor(row: dict[str, Any]) -> Professor:
    normalized = {_normalize_key(key): _cell_to_text(value) for key, value in row.items()}

    email = _pick(normalized, "email", "e-mail", "mail", "contact_email", "contact")
    if not email:
        email = _extract_email(" ".join(_cell_to_text(value) for value in row.values()))

    return Professor(
        name=_pick(normalized, "professor_name", "professor", "name", "faculty_name"),
        email=email,
        university=_pick(normalized, "iit", "university", "institute", "institution", "college"),
        department=_pick(normalized, "department", "dept", "school"),
        website=_pick(normalized, "website", "webpage", "homepage", "profile", "faculty_page"),
        lab=_pick(normalized, "lab", "laboratory", "research_lab", "group"),
        research_focus=_pick(
            normalized,
            "specialities_research_focus",
            "specialities",
            "specialties_research_focus",
            "research_focus",
            "research_interests",
            "interests",
        ),
        past_work=_pick(
            normalized,
            "past_research_key_works",
            "past_research",
            "key_works",
            "recent_papers",
            "publications",
            "research_work",
        ),
        recent_papers=_pick(
            normalized,
            "recent_papers",
            "recent_publications",
            "selected_papers",
            "selected_publications",
        ),
        alignment=_pick_alignment(normalized),
        additional_details=_pick(
            normalized,
            "additional_details",
            "details",
            "notes",
            "extra",
            "contact_details",
        ),
        openings_page=_pick(
            normalized,
            "openings_internship_page",
            "openings",
            "internship_page",
            "opportunities",
        ),
        priority=_pick(normalized, "priority", "rank", "fit_score"),
        status=_pick(normalized, "status", "outreach_status", "email_status"),
        raw=row,
    )


def _pick(row: dict[str, str], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value:
            if "email" in key or key in {"mail", "contact"}:
                found = _extract_email(value)
                return found or value
            return value
    return ""


def _pick_alignment(row: dict[str, str]) -> str:
    direct = _pick(
        row,
        "alignment_to_applicant_interests",
        "alignment_to_applicants_interests",
        "alignment",
        "fit",
        "why_match",
    )
    if direct:
        return direct

    for key, value in row.items():
        if value and key.startswith("alignment_to_") and key.endswith("_interests"):
            return value
    return ""


def _extract_email(text: str) -> str:
    match = EMAIL_RE.search(text)
    return match.group(0) if match else ""


def _normalize_key(key: Any) -> str:
    text = str(key).strip().lower()
    text = text.replace("’", "").replace("'", "")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
